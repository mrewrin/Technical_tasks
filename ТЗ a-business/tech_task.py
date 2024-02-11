import os
import logging

import pandas as pd
from fuzzywuzzy import process
import openpyxl

def process_loan_data(segment_table_path, loans_table_path):
    try:
        # Check if files exist
        if not os.path.exists(segment_table_path) or not os.path.exists(loans_table_path):
            raise FileNotFoundError("One or both of the specified files does not exist.")

        # Read data from files
        segment_credit_product_df = \
            pd.read_excel(segment_table_path, usecols=["Кредитный продукт", "Сегмент"], nrows=110)
        loans_credit_product_df = pd.read_excel(loans_table_path, usecols=["Кредитный продукт"], skiprows=5, nrows=101)

        # Convert strings to lowercase and remove spaces
        segment_credit_product_df['Кредитный продукт'] = \
            segment_credit_product_df['Кредитный продукт'].astype(str).str.lower().str.strip()
        loans_credit_product_df['Кредитный продукт'] = \
            loans_credit_product_df['Кредитный продукт'].astype(str).str.lower().str.strip()

        # Check that the required columns exist
        required_columns = ['Кредитный продукт', 'Сегмент']
        if not all(col in segment_credit_product_df.columns for col in required_columns):
            raise ValueError("Required columns are missing in the 'Сегмент.xlsx' table.")

        # Check that the "Credit Product" column contains only strings
        assert segment_credit_product_df['Кредитный продукт'].apply(lambda x: isinstance(x, str)).all(), \
            "The 'Кредитный продукт' column in the 'Сегмент.xlsx' table should contain only strings"

        # Find matching values in the "Credit Product" field using fuzzy matching
        for idx, row in loans_credit_product_df.iterrows():
            if isinstance(row['Кредитный продукт'], str):  # Check if the value is a string
                match = process.extractOne(row['Кредитный продукт'], segment_credit_product_df['Кредитный продукт'])
                if match[1] >= 90:  # Set the matching threshold
                    loans_credit_product_df.at[idx, 'Сегмент'] = segment_credit_product_df.loc[match[2], 'Сегмент']

        # Create a temporary file to save the updated data
        with pd.ExcelWriter("ActiveLoans_2024_temp.xlsx", engine="xlsxwriter") as writer:
            loans_credit_product_df.to_excel(writer, index=False)

        # Load the original file
        wb = openpyxl.load_workbook(loans_table_path)
        ws = wb["Sheet0"]

        # Insert the header "Segment" before inserting data from the temporary file
        ws.cell(row=6, column=3, value="Сегмент")
        # Highlight the header cell in bold font
        ws['C6'].font = openpyxl.styles.Font(bold=True)

        # Transfer data from the temporary file to the "Sheet0" sheet of the original file
        temp_df = pd.read_excel("ActiveLoans_2024_temp.xlsx")
        for index, row in temp_df.iterrows():
            ws.cell(row=index+7, column=3, value=row["Сегмент"])

        # Save the original file
        wb.save(loans_table_path)

        # Remove the temporary file
        os.remove("ActiveLoans_2024_temp.xlsx")

    except FileNotFoundError as fnfe:
        logging.error(f"File not found: {fnfe.filename}")
    except pd.errors.ParserError as pe:
        logging.error(f"Parser error: {pe}")
    except AssertionError as ae:
        logging.error(ae)
    except ValueError as ve:
        logging.error(ve)
    except Exception as e:
        logging.error(f"An error occurred: {e}")

# Specify file paths
segment_table_path = "Сегмент.xlsx"
loans_table_path = "ActiveLoans_2024.xlsx"

# Call the function with file paths
process_loan_data(segment_table_path, loans_table_path)
